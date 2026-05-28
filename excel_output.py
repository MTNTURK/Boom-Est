"""
EST 4 Excel Output Generator
Produces professional .xlsx estimates matching the EST 4 format.
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
import os
from datetime import datetime


# ── Color palette ──────────────────────────────────────────────────────────────
C_NAVY      = "1B2A4A"   # header backgrounds
C_DARK_GRAY = "2D2D2D"   # division headers
C_MID_GRAY  = "4A4A4A"   # sub-headers
C_LIGHT_BLU = "D6E4F7"   # alternating rows
C_WHITE     = "FFFFFF"
C_GOLD      = "C9A84C"   # accent
C_SUBTOTAL  = "E8EDF3"   # subtotal rows
C_TOTAL_BG  = "1B2A4A"   # total row background

THIN = Side(style="thin", color="CCCCCC")
MED  = Side(style="medium", color="888888")
THICK = Side(style="medium", color=C_NAVY)


def _border(left=None, right=None, top=None, bottom=None):
    return Border(left=left or Side(style=None),
                  right=right or Side(style=None),
                  top=top or Side(style=None),
                  bottom=bottom or Side(style=None))


def _cell(ws, row, col, value="", bold=False, italic=False, size=10,
          color="000000", bg=None, align="left", number_format=None,
          border=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, italic=italic, size=size, color=color)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if number_format:
        c.number_format = number_format
    if border:
        c.border = border
    return c


def generate_excel(estimate: dict, output_path: str) -> str:
    """Generate an EST 4 Excel estimate file and return the path."""
    wb = openpyxl.Workbook()

    _build_summary_sheet(wb, estimate)
    _build_detailed_sheet(wb, estimate)
    _build_scope_sheet(wb, estimate)

    # Remove default sheet if it wasn't used
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_path)
    return output_path


def _build_summary_sheet(wb, estimate):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    # Column widths
    widths = [2, 8, 42, 16, 16, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    proj = estimate.get("project", {})
    summ = estimate.get("summary", {})
    sf   = proj.get("floor_area_sf", 0) or 1

    row = 1

    # ── Title block ─────────────────────────────────────────────────────────
    ws.merge_cells(f"B{row}:G{row}")
    _cell(ws, row, 2,
          "EST 4  ·  CONSTRUCTION COST ESTIMATE",
          bold=True, size=14, color=C_WHITE, bg=C_NAVY, align="center")
    ws.row_dimensions[row].height = 28
    row += 1

    project_title = proj.get("name", "Project").upper()
    ws.merge_cells(f"B{row}:G{row}")
    _cell(ws, row, 2, project_title,
          bold=True, size=11, color=C_WHITE, bg=C_NAVY, align="center")
    ws.row_dimensions[row].height = 20
    row += 1

    addr = proj.get("address", "")
    if addr:
        ws.merge_cells(f"B{row}:G{row}")
        _cell(ws, row, 2, addr, size=9, color="AAAAAA", bg=C_NAVY, align="center")
        ws.row_dimensions[row].height = 16
        row += 1

    row += 1  # spacer

    # ── Project info ────────────────────────────────────────────────────────
    info_fields = [
        ("Project:", proj.get("name", "—")),
        ("Address:", proj.get("address", "—")),
        ("Client:", proj.get("client", "—")),
        ("Architect:", proj.get("architect", "—")),
        ("Floor Area:", f"{sf:,} SF" if sf else "—"),
        ("Estimate Date:", datetime.now().strftime("%B %Y")),
        ("Prepared by:", "EST 4 – Estimating"),
    ]
    for label, val in info_fields:
        _cell(ws, row, 2, label, bold=True, size=9, color=C_NAVY)
        ws.merge_cells(f"C{row}:G{row}")
        _cell(ws, row, 3, val, size=9)
        ws.row_dimensions[row].height = 16
        row += 1

    row += 1

    # ── Division summary header ─────────────────────────────────────────────
    ws.merge_cells(f"B{row}:G{row}")
    _cell(ws, row, 2, "ESTIMATE SUMMARY BY DIVISION",
          bold=True, size=10, color=C_WHITE, bg=C_DARK_GRAY, align="center")
    ws.row_dimensions[row].height = 18
    row += 1

    headers = ["Div", "Description", "Low ($)", "High ($)", "Mid ($)", "$/SF"]
    bg_h = C_MID_GRAY
    for i, h in enumerate(headers):
        col = i + 2
        _cell(ws, row, col, h, bold=True, size=9, color=C_WHITE, bg=bg_h,
              align="right" if i > 1 else "left")
    ws.row_dimensions[row].height = 16
    row += 1

    # Division rows
    divisions = estimate.get("divisions", [])
    total_low = total_high = 0
    alt = False
    for div in divisions:
        bg = C_LIGHT_BLU if alt else C_WHITE
        alt = not alt
        lo = div.get("subtotal_low", 0)
        hi = div.get("subtotal_high", 0)
        mid = (lo + hi) / 2
        psf = mid / sf if sf else 0
        total_low += lo
        total_high += hi

        vals = [div.get("code",""), div.get("name",""), lo, hi, mid, psf]
        fmts = [None, None, '"$"#,##0', '"$"#,##0', '"$"#,##0', '"$"#,##0.00']
        aligns = ["left","left","right","right","right","right"]
        for i, (v, fmt, al) in enumerate(zip(vals, fmts, aligns)):
            _cell(ws, row, i+2, v, size=9, bg=bg, align=al, number_format=fmt)
        ws.row_dimensions[row].height = 15
        row += 1

    # ── Subtotal ────────────────────────────────────────────────────────────
    row += 1
    mid_direct = (total_low + total_high) / 2
    _cell(ws, row, 2, "SUBTOTAL – DIRECT CONSTRUCTION COST", bold=True, size=9,
          bg=C_SUBTOTAL, align="left")
    ws.merge_cells(f"B{row}:C{row}")
    for col, val in [(4, total_low),(5, total_high),(6, mid_direct),(7, mid_direct/sf if sf else 0)]:
        _cell(ws, row, col, val, bold=True, size=9, bg=C_SUBTOTAL, align="right",
              number_format='"$"#,##0')
    ws.row_dimensions[row].height = 16
    row += 1

    gc_lo = total_low * 0.12
    gc_hi = total_high * 0.12
    gc_mid = (gc_lo + gc_hi) / 2
    _cell(ws, row, 2, "GC Overhead & Profit (12%)", size=9, bg=C_WHITE, align="left")
    ws.merge_cells(f"B{row}:C{row}")
    for col, val in [(4, gc_lo),(5, gc_hi),(6, gc_mid),(7, gc_mid/sf if sf else 0)]:
        _cell(ws, row, col, val, size=9, bg=C_WHITE, align="right", number_format='"$"#,##0')
    ws.row_dimensions[row].height = 15
    row += 1

    cont_lo = (total_low + gc_lo) * 0.05
    cont_hi = (total_high + gc_hi) * 0.05
    cont_mid = (cont_lo + cont_hi) / 2
    _cell(ws, row, 2, "Design / Bid Contingency (5%)", size=9, bg=C_LIGHT_BLU, align="left")
    ws.merge_cells(f"B{row}:C{row}")
    for col, val in [(4, cont_lo),(5, cont_hi),(6, cont_mid),(7, cont_mid/sf if sf else 0)]:
        _cell(ws, row, col, val, size=9, bg=C_LIGHT_BLU, align="right", number_format='"$"#,##0')
    ws.row_dimensions[row].height = 15
    row += 1

    grand_lo  = total_low  + gc_lo  + cont_lo
    grand_hi  = total_high + gc_hi  + cont_hi
    grand_mid = (grand_lo + grand_hi) / 2
    ws.merge_cells(f"B{row}:C{row}")
    _cell(ws, row, 2, "TOTAL PROJECT ESTIMATE", bold=True, size=10, color=C_WHITE,
          bg=C_TOTAL_BG, align="left")
    for col, val in [(4, grand_lo),(5, grand_hi),(6, grand_mid),(7, grand_mid/sf if sf else 0)]:
        _cell(ws, row, col, val, bold=True, size=10, color=C_WHITE, bg=C_TOTAL_BG,
              align="right", number_format='"$"#,##0')
    ws.row_dimensions[row].height = 20
    row += 2

    # ── Exclusions ──────────────────────────────────────────────────────────
    ws.merge_cells(f"B{row}:G{row}")
    _cell(ws, row, 2, "EXCLUSIONS / CLARIFICATIONS", bold=True, size=9,
          color=C_WHITE, bg=C_DARK_GRAY)
    ws.row_dimensions[row].height = 16
    row += 1

    for i, excl in enumerate(estimate.get("exclusions", []), 1):
        ws.merge_cells(f"B{row}:G{row}")
        _cell(ws, row, 2, f"{i}.  {excl}", size=8, wrap=True)
        ws.row_dimensions[row].height = 14
        row += 1

    # Scope summary at bottom
    if proj.get("scope_summary"):
        row += 1
        ws.merge_cells(f"B{row}:G{row}")
        _cell(ws, row, 2, "SCOPE SUMMARY", bold=True, size=9, color=C_WHITE, bg=C_NAVY)
        ws.row_dimensions[row].height = 16
        row += 1
        ws.merge_cells(f"B{row}:G{row}")
        _cell(ws, row, 2, proj["scope_summary"], size=9, wrap=True)
        ws.row_dimensions[row].height = 40


def _build_detailed_sheet(wb, estimate):
    ws = wb.create_sheet("Detailed Estimate")
    ws.sheet_view.showGridLines = False

    proj = estimate.get("project", {})
    sf   = proj.get("floor_area_sf", 0) or 1

    widths = [2, 8, 44, 6, 8, 13, 13, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    title = f"EST 4  ·  DETAILED CONSTRUCTION ESTIMATE  ·  {proj.get('name','').upper()}"
    if sf and sf != 1:
        title += f"  ·  {sf:,} SF"
    ws.merge_cells(f"B{row}:I{row}")
    _cell(ws, row, 2, title, bold=True, size=11, color=C_WHITE, bg=C_NAVY, align="center")
    ws.row_dimensions[row].height = 24
    row += 2

    col_headers = ["Ref", "Item / Description of Work", "UOM", "Qty",
                   "Unit Low ($)", "Unit High ($)", "Total Low ($)", "Total High ($)"]
    for i, h in enumerate(col_headers):
        col = i + 2
        _cell(ws, row, col, h, bold=True, size=9, color=C_WHITE, bg=C_MID_GRAY,
              align="right" if i > 2 else "left")
    ws.row_dimensions[row].height = 16
    row += 1

    divisions = estimate.get("divisions", [])
    for div in divisions:
        # Division header
        ws.merge_cells(f"B{row}:I{row}")
        _cell(ws, row, 2,
              f"  DIVISION {div.get('code','')}  ·  {div.get('name','').upper()}",
              bold=True, size=9, color=C_WHITE, bg=C_DARK_GRAY)
        ws.row_dimensions[row].height = 16
        row += 1

        alt = False
        for item in div.get("items", []):
            bg = C_LIGHT_BLU if alt else C_WHITE
            alt = not alt
            vals = [
                item.get("ref",""),
                item.get("description",""),
                item.get("uom",""),
                item.get("qty", 0),
                item.get("unit_low", 0),
                item.get("unit_high", 0),
                item.get("total_low", 0),
                item.get("total_high", 0),
            ]
            fmts = [None, None, None, "#,##0",
                    '"$"#,##0.00', '"$"#,##0.00', '"$"#,##0', '"$"#,##0']
            aligns = ["left","left","center","right","right","right","right","right"]
            for i, (v, fmt, al) in enumerate(zip(vals, fmts, aligns)):
                _cell(ws, row, i+2, v, size=9, bg=bg, align=al, number_format=fmt,
                      wrap=(i==1))
            ws.row_dimensions[row].height = 15
            row += 1

        # Subtotal row
        lo = div.get("subtotal_low", 0)
        hi = div.get("subtotal_high", 0)
        ws.merge_cells(f"B{row}:F{row}")
        _cell(ws, row, 2, "DIVISION SUBTOTAL", bold=True, size=9, bg=C_SUBTOTAL)
        _cell(ws, row, 8, lo, bold=True, size=9, bg=C_SUBTOTAL, align="right",
              number_format='"$"#,##0')
        _cell(ws, row, 9, hi, bold=True, size=9, bg=C_SUBTOTAL, align="right",
              number_format='"$"#,##0')
        ws.row_dimensions[row].height = 15
        row += 2

    # Footer note
    ws.merge_cells(f"B{row}:I{row}")
    _cell(ws, row, 2,
          "  ►  SEE SUMMARY SHEET FOR TOTAL PROJECT ESTIMATE INCLUDING GC OVERHEAD, PROFIT & CONTINGENCY",
          italic=True, size=8, color="666666")


def _build_scope_sheet(wb, estimate):
    ws = wb.create_sheet("Scope of Work")
    ws.sheet_view.showGridLines = False

    proj = estimate.get("project", {})

    widths = [2, 8, 18, 55, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    title = f"DETAILED SCOPE OF WORK  ·  {proj.get('name','').upper()}"
    ws.merge_cells(f"B{row}:F{row}")
    _cell(ws, row, 2, title, bold=True, size=11, color=C_WHITE, bg=C_NAVY, align="center")
    ws.row_dimensions[row].height = 24
    row += 2

    headers = ["Ref", "Trade / Area", "Scope Item Description", "Plans Ref"]
    for i, h in enumerate(headers):
        _cell(ws, row, i+2, h, bold=True, size=9, color=C_WHITE, bg=C_MID_GRAY)
    ws.row_dimensions[row].height = 16
    row += 1

    scope_items = estimate.get("scope_of_work", [])
    last_trade = None
    alt = False
    for item in scope_items:
        trade = item.get("trade", "")
        if trade != last_trade:
            # Trade section header
            ws.merge_cells(f"B{row}:F{row}")
            section = trade.upper() if trade else "GENERAL"
            _cell(ws, row, 2, f"  {section}", bold=True, size=9,
                  color=C_WHITE, bg=C_DARK_GRAY)
            ws.row_dimensions[row].height = 16
            row += 1
            last_trade = trade
            alt = False

        bg = C_LIGHT_BLU if alt else C_WHITE
        alt = not alt
        _cell(ws, row, 2, item.get("ref",""), size=9, bg=bg)
        _cell(ws, row, 3, item.get("trade",""), size=9, bg=bg, wrap=True)
        _cell(ws, row, 4, item.get("description",""), size=9, bg=bg, wrap=True)
        _cell(ws, row, 5, item.get("plans_ref",""), size=8, bg=bg, color="555555")
        ws.row_dimensions[row].height = 40
        row += 1
